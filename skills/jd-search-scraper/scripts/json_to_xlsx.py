#!/usr/bin/env python3
"""
json_to_xlsx.py — Convert a JSON array to a styled Excel file.
Usage: python3 json_to_xlsx.py <input.json> <output.xlsx> [header_color_hex]

Reads JSON array of objects, creates .xlsx with auto-width columns,
frozen header row, alternating row colors, and clickable URL hyperlinks.
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

infile   = sys.argv[1]
outfile  = sys.argv[2]
color    = sys.argv[3] if len(sys.argv) > 3 else 'CC0000'

items = json.load(open(infile))
if not items:
    print("No items."); sys.exit(1)

headers = list(items[0].keys())

wb = Workbook()
ws = wb.active
hfill = PatternFill(start_color=color, end_color=color, fill_type="solid")
hfont = Font(bold=True, color="FFFFFF", size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

# Header
for col, h in enumerate(headers, 1):
    c = ws.cell(1, col, h.replace('_',' ').title())
    c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal='center'); c.border = border

# Data
for idx, item in enumerate(items):
    r = idx + 2
    for col, h in enumerate(headers, 1):
        v = item.get(h, '')
        c = ws.cell(r, col, v)
        c.border = border
        if h == 'url' and v:
            c.hyperlink = str(v)
            c.font = Font(color="0563C1", underline="single")
        if idx % 2 == 0:
            c.fill = alt_fill

# Auto-width
for col, h in enumerate(headers, 1):
    max_len = len(str(h))
    for r in range(2, min(len(items)+2, 20)):
        val = str(ws.cell(r, col).value or '')
        max_len = max(max_len, min(len(val), 60))
    ws.column_dimensions[ws.cell(1, col).column_letter].width = max_len + 4

ws.freeze_panes = 'A2'
wb.save(outfile)
print(f"Saved: {outfile} ({len(items)} rows)")
